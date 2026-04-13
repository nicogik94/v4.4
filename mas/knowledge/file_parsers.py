"""Bounded parsers for operator-uploaded files.

These helpers keep raw binaries out of ProjectState. They extract a bounded,
safe representation that the upload layer can turn into knowledge items or
structured imports.
"""
from __future__ import annotations

from dataclasses import dataclass, field
import csv
import io
from pathlib import Path

from config import UPLOAD_LAYER


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".md", ".csv", ".xlsx"}
GENERIC_MEDIA_TYPES = {"", "application/octet-stream", "binary/octet-stream"}
ALLOWED_MEDIA_TYPES = {
    ".pdf": {"application/pdf"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    },
    ".txt": {"text/plain"},
    ".md": {"text/markdown", "text/plain"},
    ".csv": {"text/csv", "application/csv", "application/vnd.ms-excel"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
}


@dataclass
class ParsedChunk:
    title: str
    text: str
    source_ref: str


@dataclass
class ParsedTable:
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, str]] = field(default_factory=list)
    csv_text: str = ""
    sheet_name: str = ""
    sheet_count: int = 0


@dataclass
class ParsedUpload:
    parser_kind: str
    media_type: str
    file_kind: str
    chunks: list[ParsedChunk] = field(default_factory=list)
    table: ParsedTable | None = None
    page_count: int = 0
    row_count: int = 0
    sheet_count: int = 0
    sheet_name: str = ""


class UploadParseError(ValueError):
    """Raised when an uploaded file cannot be safely parsed."""


def validate_upload_type(filename: str, media_type: str) -> tuple[str, str]:
    extension = Path(filename or "").suffix.lower().strip()
    if extension not in SUPPORTED_EXTENSIONS:
        raise UploadParseError("Unsupported file type. Supported: PDF, DOCX, TXT, MD, CSV, XLSX.")
    normalized_media_type = (media_type or "").strip().lower()
    allowed_media_types = ALLOWED_MEDIA_TYPES.get(extension, set())
    if normalized_media_type and normalized_media_type not in GENERIC_MEDIA_TYPES and normalized_media_type not in allowed_media_types:
        raise UploadParseError(
            f"Unsupported media type for {extension or 'file'}: {normalized_media_type}"
        )
    if not normalized_media_type:
        normalized_media_type = next(iter(allowed_media_types or {"application/octet-stream"}))
    return extension, normalized_media_type


def parse_upload_bytes(
    filename: str,
    media_type: str,
    content: bytes,
    *,
    sheet_name: str = "",
) -> ParsedUpload:
    extension, normalized_media_type = validate_upload_type(filename, media_type)
    if len(content) > UPLOAD_LAYER.max_file_bytes:
        raise UploadParseError(
            f"File exceeds the {UPLOAD_LAYER.max_file_bytes:,}-byte upload limit."
        )
    if extension == ".pdf":
        return _parse_pdf(filename, normalized_media_type, content)
    if extension == ".docx":
        return _parse_docx(filename, normalized_media_type, content)
    if extension in {".txt", ".md"}:
        return _parse_text(filename, normalized_media_type, content, parser_kind=extension.lstrip("."))
    if extension == ".csv":
        return _parse_csv(filename, normalized_media_type, content)
    if extension == ".xlsx":
        return _parse_xlsx(filename, normalized_media_type, content, sheet_name=sheet_name)
    raise UploadParseError(f"Unsupported extension: {extension}")


def _parse_pdf(filename: str, media_type: str, content: bytes) -> ParsedUpload:
    try:
        from pypdf import PdfReader
    except Exception as exc:  # pragma: no cover - import availability validated in container
        raise UploadParseError("PDF parsing is unavailable because pypdf is not installed.") from exc

    try:
        reader = PdfReader(io.BytesIO(content))
        page_texts: list[str] = []
        for page in reader.pages:
            text = (page.extract_text() or "").strip()
            if text:
                page_texts.append(text)
    except Exception as exc:
        raise UploadParseError(f"Malformed or unreadable PDF: {exc}") from exc

    combined = "\n\n".join(page_texts).strip()
    chunks = _chunk_document_text(filename, combined)
    return ParsedUpload(
        parser_kind="pdf",
        media_type=media_type,
        file_kind="document",
        chunks=chunks,
        page_count=len(reader.pages),
    )


def _parse_docx(filename: str, media_type: str, content: bytes) -> ParsedUpload:
    try:
        from docx import Document
    except Exception as exc:  # pragma: no cover - dependency exists in runtime image
        raise UploadParseError("DOCX parsing is unavailable because python-docx is not installed.") from exc

    try:
        document = Document(io.BytesIO(content))
        paragraphs = [(paragraph.text or "").strip() for paragraph in document.paragraphs]
        combined = "\n".join(text for text in paragraphs if text).strip()
    except Exception as exc:
        raise UploadParseError(f"Malformed or unreadable DOCX: {exc}") from exc

    chunks = _chunk_document_text(filename, combined)
    return ParsedUpload(
        parser_kind="docx",
        media_type=media_type,
        file_kind="document",
        chunks=chunks,
    )


def _parse_text(filename: str, media_type: str, content: bytes, *, parser_kind: str) -> ParsedUpload:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("utf-8", errors="replace")
    chunks = _chunk_document_text(filename, text)
    return ParsedUpload(
        parser_kind=parser_kind,
        media_type=media_type,
        file_kind="document",
        chunks=chunks,
    )


def _parse_csv(filename: str, media_type: str, content: bytes) -> ParsedUpload:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("utf-8", errors="replace")
    table = _parse_csv_text(text, sheet_name="")
    return ParsedUpload(
        parser_kind="csv",
        media_type=media_type,
        file_kind="table",
        table=table,
        row_count=len(table.rows),
    )


def _parse_xlsx(filename: str, media_type: str, content: bytes, *, sheet_name: str = "") -> ParsedUpload:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - import availability validated in container
        raise UploadParseError("XLSX parsing is unavailable because openpyxl is not installed.") from exc

    try:
        workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise UploadParseError(f"Malformed or unreadable XLSX: {exc}") from exc

    visible_sheets = [worksheet for worksheet in workbook.worksheets if worksheet.sheet_state == "visible"]
    if not visible_sheets:
        raise UploadParseError("XLSX workbook has no visible sheets.")
    target_sheet = None
    normalized_sheet_name = (sheet_name or "").strip()
    if normalized_sheet_name:
        for worksheet in visible_sheets:
            if worksheet.title == normalized_sheet_name:
                target_sheet = worksheet
                break
        if target_sheet is None:
            raise UploadParseError(f"Sheet not found: {normalized_sheet_name}")
    else:
        target_sheet = visible_sheets[0]

    rows = list(target_sheet.iter_rows(values_only=True))
    if not rows:
        raise UploadParseError("XLSX sheet is empty.")
    headers = [_clean_header(cell, index) for index, cell in enumerate(rows[0])]
    csv_buffer = io.StringIO()
    writer = csv.writer(csv_buffer)
    writer.writerow(headers)
    normalized_rows: list[dict[str, str]] = []
    for raw_index, raw_row in enumerate(rows[1:], start=2):
        if len(normalized_rows) >= UPLOAD_LAYER.max_table_rows:
            break
        values = [_clean_cell(cell) for cell in raw_row[: len(headers)]]
        row_dict = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if not any(str(value).strip() for value in row_dict.values()):
            continue
        normalized_rows.append(row_dict)
        writer.writerow([row_dict.get(header, "") for header in headers])
    table = ParsedTable(
        headers=headers,
        rows=normalized_rows,
        csv_text=csv_buffer.getvalue(),
        sheet_name=target_sheet.title,
        sheet_count=len(visible_sheets),
    )
    return ParsedUpload(
        parser_kind="xlsx",
        media_type=media_type,
        file_kind="table",
        table=table,
        row_count=len(table.rows),
        sheet_count=len(visible_sheets),
        sheet_name=target_sheet.title,
    )


def _parse_csv_text(text: str, *, sheet_name: str) -> ParsedTable:
    try:
        reader = csv.DictReader(io.StringIO(text))
    except csv.Error as exc:
        raise UploadParseError(f"Malformed CSV header: {exc}") from exc
    if not reader.fieldnames:
        raise UploadParseError("CSV header row is missing.")

    headers = [_clean_header(value, index) for index, value in enumerate(reader.fieldnames)]
    rows: list[dict[str, str]] = []
    try:
        for row in reader:
            if len(rows) >= UPLOAD_LAYER.max_table_rows:
                break
            normalized = {header: _truncate_cell(row.get(reader.fieldnames[index], "")) for index, header in enumerate(headers)}
            if not any(value.strip() for value in normalized.values()):
                continue
            rows.append(normalized)
    except csv.Error as exc:
        raise UploadParseError(f"Malformed CSV rows: {exc}") from exc
    return ParsedTable(
        headers=headers,
        rows=rows,
        csv_text=text,
        sheet_name=sheet_name,
        sheet_count=1,
    )


def _chunk_document_text(filename: str, text: str) -> list[ParsedChunk]:
    normalized = (text or "").replace("\r\n", "\n").strip()
    if not normalized:
        raise UploadParseError("No readable text could be extracted from the uploaded file.")
    clipped = normalized[: UPLOAD_LAYER.max_document_chars]
    paragraphs = [paragraph.strip() for paragraph in clipped.split("\n\n") if paragraph.strip()]
    chunks: list[ParsedChunk] = []
    buffer = ""
    chunk_index = 1
    for paragraph in paragraphs or [clipped]:
        candidate = f"{buffer}\n\n{paragraph}".strip() if buffer else paragraph
        if buffer and len(candidate) > UPLOAD_LAYER.document_chunk_chars:
            chunks.append(_make_chunk(filename, buffer, chunk_index))
            chunk_index += 1
            buffer = paragraph
            if len(chunks) >= UPLOAD_LAYER.max_document_chunks:
                break
            continue
        if len(paragraph) > UPLOAD_LAYER.document_chunk_chars:
            for offset in range(0, len(paragraph), UPLOAD_LAYER.document_chunk_chars):
                piece = paragraph[offset : offset + UPLOAD_LAYER.document_chunk_chars].strip()
                if not piece:
                    continue
                if buffer:
                    chunks.append(_make_chunk(filename, buffer, chunk_index))
                    chunk_index += 1
                    buffer = ""
                chunks.append(_make_chunk(filename, piece, chunk_index))
                chunk_index += 1
                if len(chunks) >= UPLOAD_LAYER.max_document_chunks:
                    break
            if len(chunks) >= UPLOAD_LAYER.max_document_chunks:
                break
            continue
        buffer = candidate
    if buffer and len(chunks) < UPLOAD_LAYER.max_document_chunks:
        chunks.append(_make_chunk(filename, buffer, chunk_index))
    return chunks[: UPLOAD_LAYER.max_document_chunks]


def _make_chunk(filename: str, text: str, chunk_index: int) -> ParsedChunk:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    first_line = lines[0] if lines else Path(filename).stem
    title = f"{Path(filename).name} — {first_line[:80]}"
    return ParsedChunk(
        title=title,
        text=text[: UPLOAD_LAYER.document_chunk_chars],
        source_ref=f"{Path(filename).name}#chunk={chunk_index}",
    )


def _clean_header(value: object, index: int) -> str:
    candidate = str(value or "").strip()
    return candidate[:80] if candidate else f"column_{index + 1}"


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    return _truncate_cell(str(value))


def _truncate_cell(value: str) -> str:
    return str(value or "").strip()[: UPLOAD_LAYER.max_cell_chars]
