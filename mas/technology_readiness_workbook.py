"""XLSX workbook projection for Technology Readiness & Transfer audits."""
from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from state import ProjectState
from tools.technology_readiness import (
    EVIDENCE_CATEGORIES,
    IP_PROTECTION_AXES,
    RESEARCH_INDUSTRY_CRITERIA,
    TRL_PHASES,
    WORKBOOK_DISCLAIMER,
    build_claim_ledger,
    build_stage_gate_decision,
    build_tto_handoff_package,
    compute_alignment_score,
    compute_evidence_sufficiency,
)


TECHNOLOGY_READINESS_WORKBOOK_PROFILE = "technology_readiness_workbook"

TECHNOLOGY_READINESS_WORKBOOK_SHEETS: tuple[str, ...] = (
    "Executive Summary",
    "TRL Diagnosis",
    "Evidence Register",
    "Research-Industry Matrix",
    "IP Protection Axis",
    "Next-Level Recommendations",
    "Stage-Gate Decisions",
    "Technical Validation Plan",
    "Industrial Transfer Plan",
    "Readiness Roadmap",
    "Go-No-Go Checklist",
    "Claim Ledger",
    "TTO Handoff",
)


def technology_readiness_workbook_xlsx_bytes(state: ProjectState) -> bytes:
    """Serialize a technology-readiness project into an operator-review XLSX."""
    workbook = Workbook()
    workbook.active.title = TECHNOLOGY_READINESS_WORKBOOK_SHEETS[0]

    _write_executive_summary(workbook[TECHNOLOGY_READINESS_WORKBOOK_SHEETS[0]], state)
    _write_trl_diagnosis(workbook.create_sheet("TRL Diagnosis"), state)
    _write_evidence_register(workbook.create_sheet("Evidence Register"), state)
    _write_research_industry_matrix(workbook.create_sheet("Research-Industry Matrix"), state)
    _write_ip_axis(workbook.create_sheet("IP Protection Axis"), state)
    _write_next_level(workbook.create_sheet("Next-Level Recommendations"), state)
    _write_stage_gate(workbook.create_sheet("Stage-Gate Decisions"), state)
    _write_technical_validation(workbook.create_sheet("Technical Validation Plan"), state)
    _write_industrial_transfer(workbook.create_sheet("Industrial Transfer Plan"), state)
    _write_readiness_roadmap(workbook.create_sheet("Readiness Roadmap"), state)
    _write_go_no_go(workbook.create_sheet("Go-No-Go Checklist"), state)
    _write_claim_ledger(workbook.create_sheet("Claim Ledger"), state)
    _write_tto_handoff(workbook.create_sheet("TTO Handoff"), state)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def stage_gate_decision_for_state(state: ProjectState) -> dict[str, Any]:
    """Build the workbook/API-ready stage-gate decision from current state."""
    trl = _phase_dict(state, "trl_diagnosis")
    next_level = _phase_dict(state, "next_level_recommendations")
    summary = _phase_dict(state, "executive_summary")
    roadmap = _phase_dict(state, "readiness_roadmap")
    technical = _phase_dict(state, "technical_validation_plan")
    ip_axis = _phase_dict(state, "ip_protection_axis")
    categories = evidence_categories_for_state(state)

    return build_stage_gate_decision(
        {
            "current_trl": trl.get("current_trl") or next_level.get("current_trl") or summary.get("current_trl"),
            "next_target_trl": next_level.get("next_target_trl") or summary.get("target_trl") or trl.get("target_trl"),
            "evidence_categories": categories,
            "required_tests": next_level.get("required_tests") or _validation_test_names(technical),
            "required_evidence": next_level.get("required_evidence"),
            "acceptance_criteria": technical.get("acceptance_criteria") or next_level.get("advancement_criteria"),
            "owner_suggestions": next_level.get("suggested_owners"),
            "estimated_time_range": next_level.get("estimated_time_range"),
            "go_no_go_criteria": roadmap.get("go_no_go_criteria") or next_level.get("advancement_criteria"),
            "confidence": next_level.get("confidence") or trl.get("confidence") or summary.get("confidence"),
            "ip_claims_present": bool(ip_axis),
        }
    )


def claim_ledger_for_state(state: ProjectState) -> dict[str, Any]:
    """Build a structured claim ledger using only supplied evidence IDs."""
    trl = _phase_dict(state, "trl_diagnosis")
    next_level = _phase_dict(state, "next_level_recommendations")
    summary = _phase_dict(state, "executive_summary")
    ip_axis = _phase_dict(state, "ip_protection_axis")
    categories = evidence_categories_for_state(state)
    evidence_ids = _evidence_ids_for_categories(state, categories)
    ip_evidence_ids = _evidence_ids_for_categories(state, ["ip_review"])

    return build_claim_ledger(
        {
            "current_trl": trl.get("current_trl") or summary.get("current_trl"),
            "confidence": trl.get("confidence") or summary.get("confidence"),
            "why_not_higher": trl.get("why_not_higher"),
            "evidence_categories": categories,
            "evidence_ids": evidence_ids,
            "required_evidence": next_level.get("required_evidence"),
            "readiness_verdict": summary.get("readiness_verdict") or summary.get("readiness_verdict_code"),
            "readiness_verdict_code": summary.get("readiness_verdict_code"),
            "ip_protection_axis": ip_axis,
            "ip_evidence_ids": ip_evidence_ids,
        }
    )


def evidence_categories_for_state(state: ProjectState) -> list[str]:
    categories: list[str] = []
    for evidence in getattr(state, "imported_evidence", []) or []:
        category = str(getattr(evidence, "category", "") or "").strip()
        if category:
            categories.append(category)
    inventory = _phase_dict(state, "scientific_inventory")
    for item in inventory.get("evidence_items") or []:
        item_dict = _dict(item)
        category = str(item_dict.get("category") or "").strip()
        if category:
            categories.append(category)
    return sorted({category for category in categories if category})


def _write_executive_summary(worksheet: Any, state: ProjectState) -> None:
    trl = _phase_dict(state, "trl_diagnosis")
    summary = _phase_dict(state, "executive_summary")
    stage_gate = stage_gate_decision_for_state(state)
    rows = [
        ["Field", "Value"],
        ["Disclaimer", WORKBOOK_DISCLAIMER],
        ["Project ID", state.project_id],
        ["Project name", state.project_name],
        ["Project type", state.project_type],
        ["Current TRL", summary.get("current_trl") or trl.get("current_trl") or "Not supplied"],
        ["Target TRL", summary.get("target_trl") or trl.get("target_trl") or "Not supplied"],
        ["Confidence", summary.get("confidence") or trl.get("confidence") or "Not supplied"],
        ["Readiness verdict code", summary.get("readiness_verdict_code") or "not_assessable"],
        ["Readiness verdict", summary.get("readiness_verdict") or "Not supplied"],
        ["Why not higher", trl.get("why_not_higher") or "Not supplied"],
        ["Top blockers", _join(summary.get("top_blockers"))],
        ["Recommended next step", summary.get("recommended_next_step") or "Not supplied"],
        ["Operator summary", summary.get("operator_summary") or "Not supplied"],
        ["Stage-gate decision", stage_gate["decision"]],
    ]
    _write_rows(worksheet, rows)


def _write_trl_diagnosis(worksheet: Any, state: ProjectState) -> None:
    trl = _phase_dict(state, "trl_diagnosis")
    rows = [
        ["Field", "Value"],
        ["Current TRL", trl.get("current_trl") or "Not supplied"],
        ["Target TRL", trl.get("target_trl") or "Not supplied"],
        ["Confidence", trl.get("confidence") or "Not supplied"],
        ["Current phase name", trl.get("current_phase_name") or "Not supplied"],
        ["Evidence supporting current TRL", _join(trl.get("evidence_supporting_current_trl"))],
        ["Why not higher", trl.get("why_not_higher") or "Not supplied"],
        ["Missing evidence", _join(trl.get("evidence_gaps"))],
        ["Legal/certification boundary", trl.get("legal_or_certification_disclaimer") or WORKBOOK_DISCLAIMER],
    ]
    _write_rows(worksheet, rows)


def _write_evidence_register(worksheet: Any, state: ProjectState) -> None:
    categories = set(evidence_categories_for_state(state))
    evidence_by_category = _evidence_by_category(state)
    rows = [["Evidence category", "Status", "Evidence IDs", "Notes"]]
    for category in EVIDENCE_CATEGORIES:
        evidence_ids = evidence_by_category.get(category, [])
        status = "supplied" if category in categories or evidence_ids else "missing"
        rows.append([category, status, _join(evidence_ids), _evidence_notes(category, state)])
    unknown = compute_evidence_sufficiency(categories)["unknown_categories"]
    for category in unknown:
        rows.append([category, "unknown category", "", "Review taxonomy mapping before relying on this evidence."])
    _write_rows(worksheet, rows)


def _write_research_industry_matrix(worksheet: Any, state: ProjectState) -> None:
    alignment = _phase_dict(state, "research_industry_alignment")
    scores = _dict(alignment.get("criteria_scores"))
    rows = [["Criterion", "Score", "Evidence", "Gap", "Recommendation"]]
    for criterion in RESEARCH_INDUSTRY_CRITERIA:
        item = _dict(scores.get(criterion))
        rows.append([
            criterion,
            item.get("score", "Not supplied"),
            item.get("evidence") or "Not supplied",
            item.get("gap") or "Not supplied",
            item.get("recommendation") or "Not supplied",
        ])
    rows.append(["overall_alignment_score", alignment.get("overall_alignment_score") or compute_alignment_score(scores), "", "", ""])
    rows.append(["top_alignment_strengths", "", _join(alignment.get("top_alignment_strengths")), "", ""])
    rows.append(["top_alignment_gaps", "", "", _join(alignment.get("top_alignment_gaps")), ""])
    _write_rows(worksheet, rows)


def _write_ip_axis(worksheet: Any, state: ProjectState) -> None:
    ip_axis = _phase_dict(state, "ip_protection_axis")
    rows = [["Axis", "Preliminary assessment", "Evidence", "Gap", "Disclosure risk", "Recommended review", "Specialist review"]]
    for axis in IP_PROTECTION_AXES:
        item = _dict(ip_axis.get(axis))
        rows.append([
            axis,
            item.get("preliminary_assessment") or "Not supplied",
            _join(item.get("evidence")),
            item.get("gap") or "Not supplied",
            item.get("disclosure_risk") or "Flag disclosure risk before publication or external demos.",
            item.get("recommended_review") or "Specialist review required.",
            "Specialist review required." if ip_axis.get("specialist_review_required", True) else "Review flag was false; operator must verify.",
        ])
    rows.append(["ip_risk_notes", _join(ip_axis.get("ip_risk_notes")), "", "", "", "", "Do not claim legal patentability."])
    _write_rows(worksheet, rows)


def _write_next_level(worksheet: Any, state: ProjectState) -> None:
    next_level = _phase_dict(state, "next_level_recommendations")
    rows = [
        ["Section", "Value"],
        ["Current TRL", next_level.get("current_trl") or "Not supplied"],
        ["Next target TRL", next_level.get("next_target_trl") or "Not supplied"],
        ["Current phase name", next_level.get("current_phase_name") or "Not supplied"],
        ["Next phase name", next_level.get("next_phase_name") or "Not supplied"],
        ["Main gap to next level", next_level.get("main_gap_to_next_level") or "Not supplied"],
        ["Recommended actions", _join(next_level.get("recommended_actions"))],
        ["Required tests", _join(next_level.get("required_tests"))],
        ["Required evidence", _join(next_level.get("required_evidence"))],
        ["Expected deliverables", _join(next_level.get("expected_deliverables"))],
        ["Risks to reduce", _join(next_level.get("risks_to_reduce"))],
        ["Suggested owners", _join(next_level.get("suggested_owners"))],
        ["Estimated time range", next_level.get("estimated_time_range") or "Not supplied"],
        ["Advancement criteria", _join(next_level.get("advancement_criteria"))],
        ["Confidence", next_level.get("confidence") or "Not supplied"],
    ]
    _write_rows(worksheet, rows)


def _write_stage_gate(worksheet: Any, state: ProjectState) -> None:
    decision = stage_gate_decision_for_state(state)
    rows = [
        ["Field", "Value"],
        ["Current TRL", decision["current_trl"]],
        ["Next TRL", decision["next_trl"]],
        ["Gate name", decision["gate_name"]],
        ["Decision", decision["decision"]],
        ["Blocking gaps", _join(decision["blocking_gaps"])],
        ["Required evidence", _join(decision["required_evidence"])],
        ["Required tests", _join(decision["required_tests"])],
        ["Acceptance criteria", _join(decision["acceptance_criteria"])],
        ["Owner suggestions", _join(decision["owner_suggestions"])],
        ["Estimated time range", decision["estimated_time_range"]],
        ["Go/no-go criteria", _join(decision["go_no_go_criteria"])],
        ["Rationale", decision["rationale"]],
        ["Confidence", decision["confidence"]],
    ]
    _write_rows(worksheet, rows)


def _write_technical_validation(worksheet: Any, state: ProjectState) -> None:
    technical = _phase_dict(state, "technical_validation_plan")
    rows = [["Section", "Value"]]
    rows.extend(["Validation test", _format_item(item)] for item in technical.get("validation_tests") or [])
    rows.append(["Acceptance criteria", _join(technical.get("acceptance_criteria"))])
    rows.append(["Measurement plan", _join(technical.get("measurement_plan"))])
    rows.append(["Failure modes", _join(technical.get("failure_modes"))])
    rows.append(["Evidence to collect", _join(technical.get("evidence_to_collect"))])
    rows.append(["Confidence", technical.get("confidence") or "Not supplied"])
    _write_rows(worksheet, rows)


def _write_industrial_transfer(worksheet: Any, state: ProjectState) -> None:
    transfer = _phase_dict(state, "industrial_transfer_plan")
    rows = [
        ["Section", "Value"],
        ["Ideal industrial partner", transfer.get("ideal_industrial_partner") or "Not supplied"],
        ["Partner validation needed", _join(transfer.get("partner_validation_needed"))],
        ["Minimum transfer package", _join(transfer.get("minimum_transfer_package"))],
        ["Transfer model options", _join(transfer.get("transfer_model_options"))],
        ["Negotiation risks", _join(transfer.get("negotiation_risks"))],
        ["Evidence required before transfer", _join(transfer.get("evidence_required_before_transfer"))],
        ["Confidence", transfer.get("confidence") or "Not supplied"],
    ]
    _write_rows(worksheet, rows)


def _write_readiness_roadmap(worksheet: Any, state: ProjectState) -> None:
    roadmap = _phase_dict(state, "readiness_roadmap")
    rows = [["TRL", "Phase name", "Time range", "Objective", "Evidence needed", "Decision gate"]]
    phases = roadmap.get("roadmap_phases") or list(TRL_PHASES.values())
    for phase in phases:
        item = _dict(phase)
        rows.append([
            item.get("trl") or "Not supplied",
            item.get("phase_name") or "Not supplied",
            item.get("time_range") or "Not supplied",
            item.get("objective") or "Not supplied",
            _join(item.get("evidence_needed")),
            item.get("decision_gate") or "Not supplied",
        ])
    rows.append(["resources_needed", "", "", _join(roadmap.get("resources_needed")), "", ""])
    rows.append(["go_no_go_criteria", "", "", _join(roadmap.get("go_no_go_criteria")), "", ""])
    _write_rows(worksheet, rows)


def _write_go_no_go(worksheet: Any, state: ProjectState) -> None:
    decision = stage_gate_decision_for_state(state)
    sufficiency = compute_evidence_sufficiency(evidence_categories_for_state(state))
    rows = [
        ["Check", "Status", "Notes"],
        ["Stage-gate decision", decision["decision"], decision["rationale"]],
        ["Evidence coverage", f"{sufficiency['coverage_count']}/{sufficiency['total_categories']}", "All categories do not need to be complete, but missing decision-critical categories must be reviewed."],
        ["Required evidence supplied", "no-go" if decision["blocking_gaps"] else "go", _join(decision["blocking_gaps"]) or "No deterministic blockers found."],
        ["IP specialist review", "go" if "ip_review" in evidence_categories_for_state(state) else "conditional", "Specialist review required before legal or disclosure reliance."],
        ["No certification claim", "go", WORKBOOK_DISCLAIMER],
        ["Operator review", "required", "Workbook must be reviewed by the operator before client delivery."],
    ]
    _write_rows(worksheet, rows)


def _write_claim_ledger(worksheet: Any, state: ProjectState) -> None:
    ledger = claim_ledger_for_state(state)
    rows = [["claim_id", "claim", "label", "confidence", "evidence_ids", "limitations", "validate_with", "would_change_if", "related_phase", "related_trl"]]
    for claim in ledger["claims"]:
        rows.append([
            claim.get("claim_id"),
            claim.get("claim"),
            claim.get("label"),
            claim.get("confidence"),
            _join(claim.get("evidence_ids")),
            _join(claim.get("limitations")),
            _join(claim.get("validate_with")),
            claim.get("would_change_if"),
            claim.get("related_phase"),
            claim.get("related_trl"),
        ])
    if ledger["warnings"]:
        rows.append(["warnings", _join(ledger["warnings"]), "", "", "", "", "", "", "", ""])
    _write_rows(worksheet, rows)


def _write_tto_handoff(worksheet: Any, state: ProjectState) -> None:
    scope = _phase_dict(state, "scope")
    trl = _phase_dict(state, "trl_diagnosis")
    package = build_tto_handoff_package(
        {
            "technology_name": scope.get("technology_name") or state.project_name,
            "current_trl": trl.get("current_trl"),
            "evidence_categories": evidence_categories_for_state(state),
        }
    )
    rows = [["Section", "Value"]]
    rows.append(["invention_disclosure_draft", _join(package["invention_disclosure_draft"])])
    rows.append(["non_confidential_summary", package["non_confidential_summary"]])
    rows.append(["confidential_technical_appendix_outline", _join(package["confidential_technical_appendix_outline"])])
    rows.append(["ip_review_questions", _join(package["ip_review_questions"])])
    rows.append(["partner_validation_brief", package["partner_validation_brief"]])
    rows.append(["commercialization_route_options", _join(package["commercialization_route_options"])])
    rows.append(["evidence_checklist_before_external_disclosure", _join(package["evidence_checklist_before_external_disclosure"])])
    rows.append(["disclosure_risk_notes", _join(package["disclosure_risk_notes"])])
    _write_rows(worksheet, rows)


def _phase_dict(state: ProjectState, phase: str) -> dict[str, Any]:
    return _dict(getattr(state, phase, None))


def _dict(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    if hasattr(value, "model_dump"):
        return value.model_dump(mode="json")
    return {}


def _join(value: Any) -> str:
    if value in (None, "", [], ()):
        return "Not supplied"
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=True, sort_keys=True)
    if isinstance(value, (list, tuple, set)):
        return "; ".join(_format_item(item) for item in value) or "Not supplied"
    return str(value)


def _format_item(item: Any) -> str:
    if item is None:
        return ""
    if isinstance(item, str):
        return item
    if isinstance(item, dict):
        return json.dumps(item, ensure_ascii=True, sort_keys=True)
    if hasattr(item, "model_dump"):
        return json.dumps(item.model_dump(mode="json"), ensure_ascii=True, sort_keys=True)
    return str(item)


def _cell(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return value
    text = _join(value) if isinstance(value, (dict, list, tuple, set)) else str(value or "")
    text = text.replace("\x00", "").strip()
    if text and text[0] in ("=", "+", "-", "@"):
        text = "'" + text
    return text


def _write_rows(worksheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        worksheet.append([_cell(value) for value in row])
    _style_sheet(worksheet)


def _style_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = wrap
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for column in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 14), 54)


def _validation_test_names(technical: dict[str, Any]) -> list[str]:
    names: list[str] = []
    for item in technical.get("validation_tests") or []:
        item_dict = _dict(item)
        names.append(str(item_dict.get("test") or item_dict.get("name") or _format_item(item)).strip())
    return [name for name in names if name]


def _evidence_by_category(state: ProjectState) -> dict[str, list[str]]:
    by_category: dict[str, list[str]] = {}
    for evidence in getattr(state, "imported_evidence", []) or []:
        category = str(getattr(evidence, "category", "") or "").strip()
        evidence_id = str(getattr(evidence, "evidence_id", "") or "").strip()
        if category and evidence_id:
            by_category.setdefault(category, []).append(evidence_id)
    return {category: sorted(set(ids)) for category, ids in by_category.items()}


def _evidence_ids_for_categories(state: ProjectState, categories: list[str] | set[str]) -> list[str]:
    wanted = set(categories or [])
    ids: list[str] = []
    for evidence in getattr(state, "imported_evidence", []) or []:
        category = str(getattr(evidence, "category", "") or "").strip()
        evidence_id = str(getattr(evidence, "evidence_id", "") or "").strip()
        if evidence_id and (not wanted or category in wanted):
            ids.append(evidence_id)
    return sorted(set(ids))


def _evidence_notes(category: str, state: ProjectState) -> str:
    notes: list[str] = []
    for evidence in getattr(state, "imported_evidence", []) or []:
        if str(getattr(evidence, "category", "") or "").strip() == category:
            title = str(getattr(evidence, "title", "") or "").strip()
            summary = str(getattr(evidence, "summary", "") or "").strip()
            note = " - ".join(part for part in (title, summary) if part)
            if note:
                notes.append(note)
    return "; ".join(notes) if notes else "Not supplied"
