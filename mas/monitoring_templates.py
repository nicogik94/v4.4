"""Spreadsheet-safe monitoring template projections.

The helpers here are read-only projections over ProjectState. They do not
create decision gates, mutate report content, or change workflow state.
"""
from __future__ import annotations

from dataclasses import dataclass
import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from report_quality import assess_report_quality_context, evidence_maturity_projection


OPERATOR_TO_DEFINE = "Operator to define"
VALIDATION_REQUIRED = "To be confirmed"
EVIDENCE_SOURCE_UNAVAILABLE = "Not supplied"
THRESHOLD_NOT_CONFIRMED = "Threshold not yet confirmed"
THRESHOLD_VALIDATION_PENDING = "To be confirmed"
SHEET_NAME = "Monitoring Template"
README_SHEET_NAME = "README"
STOP_CHANGE_SHEET_NAME = "Stop - Change Criteria"
OODA_SHEET_NAME = "OODA Schedule"
CIRCUIT_BREAKER_SHEET_NAME = "Circuit Breakers"
CANARIES_SHEET_NAME = "Canaries"
REENTRY_RISK_SHEET_NAME = "Re-entry Watch - Risks"
REVIEW_LOG_SHEET_NAME = "Review Log"
METADATA_SHEET_NAME = "Metadata - Evidence Maturity"

CLIENT_MONITORING_TEMPLATE_HEADERS: tuple[str, ...] = (
    "Metric / signal",
    "Decision or hypothesis validated",
    "Owner / role",
    "Cadence",
    "Source / evidence source",
    "Target / good sign",
    "Warning sign",
    "Stop/change-course threshold",
    "Action if triggered",
    "Evidence maturity / validation status",
    "Notes",
)

OPERATOR_TRACE_HEADERS: tuple[str, ...] = (
    "Row source",
    "Hypothesis IDs",
    "Evidence IDs",
    "Internal source refs",
    "Diagnostic notes",
)

OPERATOR_MONITORING_TEMPLATE_HEADERS: tuple[str, ...] = (
    *CLIENT_MONITORING_TEMPLATE_HEADERS,
    *OPERATOR_TRACE_HEADERS,
)

REVIEW_LOG_HEADERS: tuple[str, ...] = (
    "Review date",
    "Reviewer",
    "Review decision / status",
    "Hypothesis / experiment",
    "Signal / metric",
    "Last observed value",
    "Observation date",
    "Evidence / note",
    "Change made",
    "Follow-up owner",
    "Next review date",
    "Follow-up due date",
    "Notes",
)


@dataclass(frozen=True)
class MonitoringTemplateRow:
    metric_signal: str = OPERATOR_TO_DEFINE
    decision_or_hypothesis: str = OPERATOR_TO_DEFINE
    owner_role: str = OPERATOR_TO_DEFINE
    cadence: str = OPERATOR_TO_DEFINE
    source_evidence_source: str = EVIDENCE_SOURCE_UNAVAILABLE
    target_good_sign: str = VALIDATION_REQUIRED
    warning_sign: str = VALIDATION_REQUIRED
    stop_change_threshold: str = THRESHOLD_NOT_CONFIRMED
    action_if_triggered: str = "Review with decision owner before changing course."
    evidence_maturity_validation_status: str = VALIDATION_REQUIRED
    notes: str = ""
    row_source: str = ""
    hypothesis_ids: tuple[str, ...] = ()
    evidence_ids: tuple[str, ...] = ()
    internal_source_refs: tuple[str, ...] = ()
    diagnostic_notes: str = ""


def build_monitoring_template_rows(state: Any) -> tuple[MonitoringTemplateRow, ...]:
    """Build deterministic monitoring rows from existing report/strategy/monitor state."""
    status = _evidence_status(state)
    rows: list[MonitoringTemplateRow] = []
    rows.extend(_decision_gate_rows(state, status))
    rows.extend(_ooda_rows(state, status))
    rows.extend(_circuit_breaker_rows(state, status))
    rows.extend(_canary_rows(state, status))
    rows.extend(_strategy_metric_rows(state, status))
    rows.extend(_needs_monitoring_rows(state, status))
    if not rows:
        rows.append(
            MonitoringTemplateRow(
                evidence_maturity_validation_status=status,
                notes="Monitoring template placeholder; operator should define tracking before relying on it.",
                row_source="template_placeholder",
                diagnostic_notes="No monitor, strategy, or clear Decision Gates state was available.",
            )
        )
    return tuple(rows)


def monitoring_template_xlsx_bytes(state: Any, *, audience: str) -> bytes:
    """Serialize monitoring rows as a spreadsheet-safe XLSX workbook."""
    mode = "operator" if str(audience or "").lower() == "operator" else "client"
    headers = OPERATOR_MONITORING_TEMPLATE_HEADERS if mode == "operator" else CLIENT_MONITORING_TEMPLATE_HEADERS
    rows = build_monitoring_template_rows(state)
    workbook = Workbook()
    readme = workbook.active
    readme.title = README_SHEET_NAME
    _write_readme_sheet(readme, mode)
    _write_tabular_sheet(
        workbook.create_sheet(SHEET_NAME),
        headers,
        [_render_row(row, headers, audience=mode) for row in rows],
    )
    if mode == "client":
        _write_tabular_sheet(
            workbook.create_sheet(STOP_CHANGE_SHEET_NAME),
            headers,
            [_render_row(row, headers, audience=mode) for row in _stop_change_rows(rows)],
        )
        _write_tabular_sheet(
            workbook.create_sheet(CANARIES_SHEET_NAME),
            headers,
            [_render_row(row, headers, audience=mode) for row in _rows_by_source(rows, "monitor_canary")],
        )
    else:
        _write_tabular_sheet(
            workbook.create_sheet(OODA_SHEET_NAME),
            headers,
            [_render_row(row, headers, audience=mode) for row in _rows_with_source_prefix(rows, "monitor_ooda_")],
        )
        _write_tabular_sheet(
            workbook.create_sheet(CIRCUIT_BREAKER_SHEET_NAME),
            headers,
            [_render_row(row, headers, audience=mode) for row in _rows_by_source(rows, "monitor_circuit_breaker")],
        )
        _write_tabular_sheet(
            workbook.create_sheet(CANARIES_SHEET_NAME),
            headers,
            [_render_row(row, headers, audience=mode) for row in _rows_by_source(rows, "monitor_canary")],
        )
        _write_tabular_sheet(
            workbook.create_sheet(REENTRY_RISK_SHEET_NAME),
            headers,
            [_render_row(row, headers, audience=mode) for row in _reentry_risk_rows(rows)],
        )
        _write_metadata_sheet(workbook.create_sheet(METADATA_SHEET_NAME), state)
    _write_review_log_sheet(workbook.create_sheet(REVIEW_LOG_SHEET_NAME))
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_readme_sheet(worksheet: Any, audience: str) -> None:
    worksheet.append(["Monitoring workbook"])
    worksheet.append(["Audience", "Operator" if audience == "operator" else "Client"])
    worksheet.append(["Human review", "Human review required before client delivery."])
    worksheet.append(["Validation boundary", "Validate monitoring thresholds before using them as change-course gates."])
    worksheet.append(["Source", "Rows are exported from the existing monitoring plan, strategy outputs, and Decision Gates."])
    worksheet.append(["Automation boundary", "Monitoring rows are human-review controls; they do not trigger autonomous actions."])
    if audience == "operator":
        worksheet.append(["Operator note", "Operator workbook may include trace columns for review and troubleshooting."])
    else:
        worksheet.append(["Client note", "Client workbook omits raw internal IDs and diagnostic-only references where possible."])
    _style_sheet(worksheet, freeze=False)


def _write_tabular_sheet(worksheet: Any, headers: tuple[str, ...], rows: list[list[str]]) -> None:
    worksheet.append(list(headers))
    if rows:
        for row in rows:
            worksheet.append(row)
    else:
        worksheet.append(["Not supplied", *["Not applicable"] * (len(headers) - 1)])
    _style_sheet(worksheet, freeze=True)


def _write_review_log_sheet(worksheet: Any) -> None:
    _write_tabular_sheet(
        worksheet,
        REVIEW_LOG_HEADERS,
        [[
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "To be confirmed",
            "Record review decisions, threshold changes, and owner follow-up here.",
        ]],
    )


def _write_metadata_sheet(worksheet: Any, state: Any) -> None:
    context = assess_report_quality_context(state)
    projection = evidence_maturity_projection(state, context)
    rows = [
        ["Field", "Value"],
        ["Evidence maturity", _text(getattr(projection, "maturity", "")) or "To be confirmed"],
        ["Client-use status", _text(getattr(projection, "client_use_status", "")) or "To be confirmed"],
        ["Validation required", _text(getattr(projection, "validation_required", "")) or "To be confirmed"],
        ["Uploaded files", _text(getattr(projection, "uploaded_files", "")) or "0"],
        ["Imported evidence", _text(getattr(projection, "imported_evidence", "")) or "0"],
        ["Imported signals", _text(getattr(projection, "imported_signals", "")) or "0"],
        ["Human review", "Human review required before client delivery."],
    ]
    for row in rows:
        worksheet.append(row)
    _style_sheet(worksheet, freeze=True)


def _style_sheet(worksheet: Any, *, freeze: bool) -> None:
    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    if freeze:
        worksheet.freeze_panes = "A2"
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = wrap
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for column in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 14), 48)


def _rows_by_source(rows: tuple[MonitoringTemplateRow, ...], source: str) -> list[MonitoringTemplateRow]:
    return [row for row in rows if row.row_source == source]


def _rows_with_source_prefix(rows: tuple[MonitoringTemplateRow, ...], prefix: str) -> list[MonitoringTemplateRow]:
    return [row for row in rows if row.row_source.startswith(prefix)]


def _stop_change_rows(rows: tuple[MonitoringTemplateRow, ...]) -> list[MonitoringTemplateRow]:
    return [
        row
        for row in rows
        if row.row_source in {"decision_gates", "monitor_circuit_breaker"}
        or _text(row.stop_change_threshold) not in {"", THRESHOLD_NOT_CONFIRMED, THRESHOLD_VALIDATION_PENDING}
    ]


def _reentry_risk_rows(rows: tuple[MonitoringTemplateRow, ...]) -> list[MonitoringTemplateRow]:
    return [
        row
        for row in rows
        if row.row_source in {"strategy_preliminary_verdict", "strategy_success_metric", "decision_gates"}
    ]


def monitoring_template_cell_rows(state: Any, *, audience: str) -> tuple[tuple[str, ...], ...]:
    """Return deterministic workbook cell values for tests and preview tooling."""
    mode = "operator" if str(audience or "").lower() == "operator" else "client"
    headers = OPERATOR_MONITORING_TEMPLATE_HEADERS if mode == "operator" else CLIENT_MONITORING_TEMPLATE_HEADERS
    rendered = [tuple(headers)]
    for row in build_monitoring_template_rows(state):
        rendered.append(tuple(_render_row(row, headers, audience=mode)))
    return tuple(rendered)


def _decision_gate_rows(state: Any, status: str) -> list[MonitoringTemplateRow]:
    section = _extract_heading_section(getattr(state, "report", ""), "Decision Gates")
    if not section:
        return []
    table_rows = _parse_decision_gate_tables(section, status)
    if table_rows:
        return table_rows
    return [
        MonitoringTemplateRow(
            metric_signal="Decision Gate",
            decision_or_hypothesis="Decision Gates",
            source_evidence_source="Report Decision Gates section",
            evidence_maturity_validation_status=status,
            notes="Decision Gates section present; threshold details require operator review.",
            row_source="decision_gates",
            diagnostic_notes="Decision Gates section was present but not parsed into a clear gate table.",
        )
    ]


def _parse_decision_gate_tables(section: str, status: str) -> list[MonitoringTemplateRow]:
    rows: list[MonitoringTemplateRow] = []
    for block in _markdown_table_blocks(section):
        if len(block) < 3:
            continue
        headers = _split_table_row(block[0])
        if not headers or not _looks_like_separator(block[1]):
            continue
        normalized = [_normalize_label(header) for header in headers]
        if not _is_clear_decision_gate_header(normalized):
            continue
        for raw in block[2:]:
            values = _split_table_row(raw)
            if not any(value.strip() for value in values):
                continue
            cell = {
                normalized[index]: values[index] if index < len(values) else ""
                for index in range(len(normalized))
            }
            metric = _first_value(cell, ("signal", "metric", "gate", "decision", "condition", "criterion")) or _nth(values, 0)
            rows.append(
                MonitoringTemplateRow(
                    metric_signal=_clean_placeholder(metric, OPERATOR_TO_DEFINE),
                    decision_or_hypothesis=_clean_placeholder(
                        _first_value(cell, ("hypothesis", "validates", "decision", "gate")),
                        "Decision Gates",
                    ),
                    owner_role=_clean_placeholder(_first_value(cell, ("owner", "role")), OPERATOR_TO_DEFINE),
                    cadence=_clean_placeholder(_first_value(cell, ("cadence", "review", "timeframe", "timeline")), OPERATOR_TO_DEFINE),
                    source_evidence_source=_clean_placeholder(_first_value(cell, ("source", "evidence")), EVIDENCE_SOURCE_UNAVAILABLE),
                    target_good_sign=_clean_placeholder(_first_value(cell, ("good", "target", "proceed", "success")), VALIDATION_REQUIRED),
                    warning_sign=_clean_placeholder(_first_value(cell, ("warning", "extend", "escalate")), VALIDATION_REQUIRED),
                    stop_change_threshold=_clean_placeholder(
                        _first_value(cell, ("stop", "change", "threshold", "kill", "defer")),
                        THRESHOLD_NOT_CONFIRMED,
                    ),
                    action_if_triggered=_clean_placeholder(_first_value(cell, ("action", "triggered", "next")), "Review with decision owner before changing course."),
                    evidence_maturity_validation_status=status,
                    notes=_clean_placeholder(_first_value(cell, ("note", "caveat", "status")), ""),
                    row_source="decision_gates",
                    hypothesis_ids=tuple(_extract_ids(" ".join(values), prefix="H")),
                    evidence_ids=tuple(_extract_evidence_ids(" ".join(values))),
                    internal_source_refs=tuple(_extract_source_refs(" ".join(values))),
                )
            )
    return rows


def _ooda_rows(state: Any, status: str) -> list[MonitoringTemplateRow]:
    monitor = getattr(state, "monitor", None)
    if not monitor:
        return []
    rows: list[MonitoringTemplateRow] = []
    schedule = getattr(monitor, "ooda_schedule", None)
    for cadence, items in (
        ("Daily", getattr(schedule, "daily", []) if schedule else []),
        ("Weekly", getattr(schedule, "weekly", []) if schedule else []),
        ("Monthly", getattr(schedule, "monthly", []) if schedule else []),
    ):
        for index, item in enumerate(items or []):
            source = _text(getattr(item, "source", ""))
            metric = _text(getattr(item, "metric", ""))
            row_context = " ".join([metric, source, cadence])
            direction = _desired_direction(row_context)
            concrete_values = _extract_concrete_monitoring_values(row_context)
            rows.append(
                MonitoringTemplateRow(
                    metric_signal=_clean_placeholder(metric, OPERATOR_TO_DEFINE),
                    decision_or_hypothesis="OODA monitoring checkpoint",
                    owner_role=_clean_placeholder(getattr(item, "owner", ""), OPERATOR_TO_DEFINE),
                    cadence=cadence,
                    source_evidence_source=_clean_placeholder(source, EVIDENCE_SOURCE_UNAVAILABLE),
                    target_good_sign=_expected_trend_text(direction) if direction else VALIDATION_REQUIRED,
                    warning_sign=_movement_against_direction_text(direction) if direction else VALIDATION_REQUIRED,
                    stop_change_threshold=_concrete_threshold_or_placeholder(
                        concrete_values,
                        THRESHOLD_NOT_CONFIRMED,
                    ),
                    action_if_triggered="Review metric trend with the owner and compare against Decision Gates.",
                    evidence_maturity_validation_status=status,
                    notes="OODA schedule item; implementation control, not a separate decision gate.",
                    row_source=f"monitor_ooda_{cadence.lower()}",
                    internal_source_refs=tuple(_extract_source_refs(source)),
                    diagnostic_notes=f"monitor.ooda_schedule.{cadence.lower()}[{index}]",
                )
            )
    return rows


def _circuit_breaker_rows(state: Any, status: str) -> list[MonitoringTemplateRow]:
    monitor = getattr(state, "monitor", None)
    if not monitor:
        return []
    rows: list[MonitoringTemplateRow] = []
    for index, item in enumerate(getattr(monitor, "circuit_breakers", []) or []):
        trip = _text(getattr(item, "trip", ""))
        reset = _text(getattr(item, "reset", ""))
        strategy_ref = _text(getattr(item, "strategy_ref", ""))
        rows.append(
            MonitoringTemplateRow(
                metric_signal=_clean_placeholder(strategy_ref, "Circuit breaker"),
                decision_or_hypothesis=_clean_placeholder(strategy_ref, "Strategy control"),
                source_evidence_source="Monitoring plan",
                target_good_sign=_clean_placeholder(reset, VALIDATION_REQUIRED),
                warning_sign=_clean_placeholder(trip, VALIDATION_REQUIRED),
                stop_change_threshold=_clean_placeholder(trip, THRESHOLD_NOT_CONFIRMED),
                action_if_triggered="Pause or change course only after operator review confirms the trip condition.",
                evidence_maturity_validation_status=status,
                notes="Circuit breaker is an implementation control; Decision Gates remain source of truth.",
                row_source="monitor_circuit_breaker",
                hypothesis_ids=tuple(_extract_ids(strategy_ref, prefix="H")),
                internal_source_refs=tuple(_extract_source_refs(" ".join([trip, reset, strategy_ref]))),
                diagnostic_notes=f"monitor.circuit_breakers[{index}]",
            )
        )
    return rows


def _canary_rows(state: Any, status: str) -> list[MonitoringTemplateRow]:
    monitor = getattr(state, "monitor", None)
    if not monitor:
        return []
    rows: list[MonitoringTemplateRow] = []
    for index, item in enumerate(getattr(monitor, "canaries", []) or []):
        signal = _text(getattr(item, "signal", ""))
        direction = _text(getattr(item, "direction", ""))
        window = _text(getattr(item, "window", ""))
        meaning = _text(getattr(item, "meaning", ""))
        inferred_direction = _desired_direction(" ".join([signal, meaning, direction]))
        concrete_values = _extract_concrete_monitoring_values(" ".join([signal, direction, window, meaning]))
        rows.append(
            MonitoringTemplateRow(
                metric_signal=_clean_placeholder(signal, OPERATOR_TO_DEFINE),
                decision_or_hypothesis="Canary signal",
                cadence=_clean_placeholder(window, OPERATOR_TO_DEFINE),
                source_evidence_source=EVIDENCE_SOURCE_UNAVAILABLE,
                target_good_sign=_clean_placeholder(
                    _canary_target_text(signal=signal, direction=inferred_direction, meaning=meaning),
                    VALIDATION_REQUIRED,
                ),
                warning_sign=_movement_against_direction_text(inferred_direction)
                if inferred_direction
                else "No movement or movement against expected direction.",
                stop_change_threshold=_canary_threshold_text(window, inferred_direction, concrete_values),
                action_if_triggered="Investigate canary, review the evidence source, and compare against Decision Gates.",
                evidence_maturity_validation_status=status,
                notes="Canary is an early-warning control; threshold requires operator confirmation.",
                row_source="monitor_canary",
                internal_source_refs=tuple(_extract_source_refs(" ".join([signal, meaning]))),
                diagnostic_notes=f"monitor.canaries[{index}]",
            )
        )
    return rows


def _strategy_metric_rows(state: Any, status: str) -> list[MonitoringTemplateRow]:
    strategy = getattr(state, "strategy", None)
    if not strategy:
        return []
    rows: list[MonitoringTemplateRow] = []
    review_date = _text(getattr(strategy, "review_date", ""))
    for index, metric in enumerate(getattr(strategy, "success_metrics", []) or []):
        metric_text = _text(metric)
        direction = _desired_direction(metric_text)
        concrete_values = _extract_concrete_monitoring_values(" ".join([metric_text, review_date]))
        rows.append(
            MonitoringTemplateRow(
                metric_signal=_clean_placeholder(metric_text, OPERATOR_TO_DEFINE),
                decision_or_hypothesis="Strategy success metric",
                cadence=_clean_placeholder(review_date, OPERATOR_TO_DEFINE),
                source_evidence_source="Strategy success metrics",
                target_good_sign=_expected_trend_text(direction) if direction else VALIDATION_REQUIRED,
                warning_sign=_movement_against_direction_text(direction) if direction else VALIDATION_REQUIRED,
                stop_change_threshold=_concrete_threshold_or_placeholder(
                    concrete_values,
                    THRESHOLD_NOT_CONFIRMED,
                ),
                action_if_triggered="Review with the decision owner at the next checkpoint.",
                evidence_maturity_validation_status=status,
                notes="Success metric requires validation before it is treated as an approved gate.",
                row_source="strategy_success_metric",
                internal_source_refs=tuple(_extract_source_refs(_text(metric))),
                diagnostic_notes=f"strategy.success_metrics[{index}]",
            )
        )
    return rows


def _needs_monitoring_rows(state: Any, status: str) -> list[MonitoringTemplateRow]:
    strategy = getattr(state, "strategy", None)
    if not strategy:
        return []
    hypotheses = {
        _text(getattr(hypothesis, "id", "")): hypothesis
        for hypothesis in (getattr(state, "hypotheses", []) or [])
        if _text(getattr(hypothesis, "id", ""))
    }
    rows: list[MonitoringTemplateRow] = []
    for index, verdict in enumerate(getattr(strategy, "preliminary_verdicts", []) or []):
        verdict_value = _enum_value(getattr(verdict, "verdict", ""))
        if verdict_value != "NEEDS_MONITORING":
            continue
        hypothesis_id = _text(getattr(verdict, "id", ""))
        hypothesis = hypotheses.get(hypothesis_id)
        metric = _text(getattr(hypothesis, "signal", "")) or _text(getattr(verdict, "monitoring_plan", "")) or OPERATOR_TO_DEFINE
        confirm = _text(getattr(hypothesis, "confirm", ""))
        reject = _text(getattr(hypothesis, "reject", ""))
        monitoring_plan = _text(getattr(verdict, "monitoring_plan", ""))
        direction = _desired_direction(" ".join([metric, confirm, reject, monitoring_plan]))
        concrete_values = _extract_concrete_monitoring_values(" ".join([metric, confirm, reject, monitoring_plan]))
        evidence_ids = tuple(_text(item) for item in getattr(hypothesis, "evidence_ids", []) or [] if _text(item))
        rows.append(
            MonitoringTemplateRow(
                metric_signal=metric,
                decision_or_hypothesis=hypothesis_id or "Hypothesis needing monitoring",
                source_evidence_source="Strategy verdict",
                target_good_sign=_clean_placeholder(confirm, _expected_trend_text(direction) if direction else VALIDATION_REQUIRED),
                warning_sign=_clean_placeholder(reject, _movement_against_direction_text(direction) if direction else VALIDATION_REQUIRED),
                stop_change_threshold=_concrete_threshold_or_placeholder(
                    _extract_concrete_monitoring_values(reject) or concrete_values,
                    THRESHOLD_NOT_CONFIRMED,
                ),
                action_if_triggered="Validate with the decision owner before changing the recommendation.",
                evidence_maturity_validation_status=status,
                notes=_clean_placeholder(monitoring_plan, "Hypothesis requires monitoring."),
                row_source="strategy_preliminary_verdict",
                hypothesis_ids=(hypothesis_id,) if hypothesis_id else (),
                evidence_ids=evidence_ids,
                internal_source_refs=tuple(_extract_source_refs(" ".join([metric, _text(getattr(verdict, "evidence", ""))]))),
                diagnostic_notes=f"strategy.preliminary_verdicts[{index}]",
            )
        )
    return rows


def _evidence_status(state: Any) -> str:
    context = assess_report_quality_context(state)
    projection = evidence_maturity_projection(state, context)
    maturity = _text(getattr(projection, "maturity", "")) or VALIDATION_REQUIRED
    validation = _text(getattr(projection, "validation_required", "")) or VALIDATION_REQUIRED
    return f"{maturity} - {validation}"


def _render_row(row: MonitoringTemplateRow, headers: tuple[str, ...], *, audience: str) -> list[str]:
    values = {
        "Metric / signal": row.metric_signal,
        "Decision or hypothesis validated": row.decision_or_hypothesis,
        "Owner / role": row.owner_role,
        "Cadence": row.cadence,
        "Source / evidence source": row.source_evidence_source,
        "Target / good sign": row.target_good_sign,
        "Warning sign": row.warning_sign,
        "Stop/change-course threshold": row.stop_change_threshold,
        "Action if triggered": row.action_if_triggered,
        "Evidence maturity / validation status": row.evidence_maturity_validation_status,
        "Notes": row.notes,
        "Row source": row.row_source,
        "Hypothesis IDs": ", ".join(row.hypothesis_ids),
        "Evidence IDs": ", ".join(row.evidence_ids),
        "Internal source refs": ", ".join(row.internal_source_refs),
        "Diagnostic notes": row.diagnostic_notes,
    }
    if audience == "client":
        values = _client_enhance_row_values(values, row)
    values = _monitoring_safe_render_values(values, audience=audience)
    return [_spreadsheet_safe_cell(values.get(header, ""), audience=audience, header=header) for header in headers]


def _monitoring_safe_render_values(values: dict[str, str], *, audience: str) -> dict[str, str]:
    safe = dict(values)
    context = " ".join(
        _text(safe.get(header, ""))
        for header in (
            "Metric / signal",
            "Decision or hypothesis validated",
            "Source / evidence source",
            "Target / good sign",
            "Warning sign",
            "Cadence",
            "Notes",
        )
    )
    direction = _desired_direction(context)
    concrete_values = _extract_concrete_monitoring_values(context)

    for header, value in list(safe.items()):
        if _text(value) == OPERATOR_TO_DEFINE:
            safe[header] = _operator_to_define_fallback(header, audience=audience)
        elif OPERATOR_TO_DEFINE in _text(value):
            safe[header] = _text(value).replace(OPERATOR_TO_DEFINE, _operator_to_define_fallback(header, audience=audience))

    threshold = _text(safe.get("Stop/change-course threshold", ""))
    if threshold == THRESHOLD_NOT_CONFIRMED:
        safe["Stop/change-course threshold"] = _threshold_render_fallback(concrete_values, direction)
    elif THRESHOLD_NOT_CONFIRMED in threshold:
        safe["Stop/change-course threshold"] = threshold.replace(
            THRESHOLD_NOT_CONFIRMED,
            _threshold_render_fallback(concrete_values, direction),
        )
    return safe


def _operator_to_define_fallback(header: str, *, audience: str) -> str:
    if header == "Source / evidence source":
        return "Not supplied"
    return "To be confirmed"


def _threshold_render_fallback(concrete_values: tuple[str, ...], direction: str) -> str:
    if concrete_values:
        return _concrete_threshold_or_placeholder(concrete_values, "")
    if direction:
        return _movement_against_direction_text(direction)
    return THRESHOLD_VALIDATION_PENDING


def _client_enhance_row_values(values: dict[str, str], row: MonitoringTemplateRow) -> dict[str, str]:
    enhanced = dict(values)
    if row.row_source == "template_placeholder":
        return enhanced

    context = " ".join(
        _text(item)
        for item in (
            row.metric_signal,
            row.decision_or_hypothesis,
            row.owner_role,
            row.cadence,
            row.source_evidence_source,
            row.target_good_sign,
            row.warning_sign,
            row.stop_change_threshold,
            row.action_if_triggered,
            row.notes,
        )
    )
    concrete_values = _extract_concrete_monitoring_values(context)
    direction = _desired_direction(context)

    enhanced["Metric / signal"] = _client_hypothesis_label(enhanced["Metric / signal"], row)
    enhanced["Decision or hypothesis validated"] = _client_hypothesis_label(
        enhanced["Decision or hypothesis validated"],
        row,
    )

    if not concrete_values and not direction:
        return enhanced

    if concrete_values and enhanced["Owner / role"] == OPERATOR_TO_DEFINE:
        enhanced["Owner / role"] = "To be confirmed"
    if concrete_values and enhanced["Cadence"] == OPERATOR_TO_DEFINE:
        enhanced["Cadence"] = _first_cadence_value(concrete_values) or "To be confirmed"
    if enhanced["Target / good sign"] == VALIDATION_REQUIRED:
        enhanced["Target / good sign"] = _expected_trend_text(direction) if direction else "Review against Decision Gates"
    if enhanced["Warning sign"] == VALIDATION_REQUIRED:
        enhanced["Warning sign"] = (
            _movement_against_direction_text(direction)
            if direction
            else "Movement against the expected Decision Gate trend"
        )
    if concrete_values and enhanced["Stop/change-course threshold"] == THRESHOLD_NOT_CONFIRMED:
        enhanced["Stop/change-course threshold"] = _concrete_threshold_or_placeholder(concrete_values, "")
    if direction and "Expected trend under remediation" not in enhanced["Notes"]:
        trend = _expected_trend_text(direction)
        enhanced["Notes"] = "; ".join(part for part in (enhanced["Notes"], trend) if part)
    return enhanced


def _client_hypothesis_label(value: str, row: MonitoringTemplateRow) -> str:
    text = _text(value)
    if not row.hypothesis_ids and not re.search(r"\bH\d+\b", text, re.I):
        return text
    raw_ids = tuple(row.hypothesis_ids) or tuple(_extract_ids(text, prefix="H"))
    if not raw_ids:
        return text

    def repl(match: re.Match[str]) -> str:
        return f"hypothesis {match.group(1)}"

    readable = re.sub(r"\bH(\d+)\b", repl, text, flags=re.I)
    if re.fullmatch(r"hypothesis\s+\d+", readable, re.I):
        label_source = _text(row.metric_signal)
        label_source = re.sub(r"\bH\d+\b", "", label_source, flags=re.I).strip(" -:;")
        if label_source and _normalize_label(label_source) != _normalize_label(text):
            readable = f"{readable} - {_short_hypothesis_topic(label_source)}"
    return readable


def _short_hypothesis_topic(value: str) -> str:
    text = _text(value)
    text = re.sub(r"\b(?:increase|decrease|improve|reduce|watch|track|monitor|validate)\b", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" -:;")
    return text[:80].strip() or "monitoring signal"


def _desired_direction(value: str) -> str:
    text = _normalize_label(value)
    if not text:
        return ""
    inverse_patterns = (
        r"\btime\s+to\s+value\b",
        r"\btime\s+value\b",
        r"\blag\b",
        r"\blatency\b",
        r"\bdelay\b",
        r"\bcycle\s+time\b",
        r"\bresponse\s+time\b",
        r"\bresolution\s+time\b",
        r"\bchurn\b",
        r"\bfailure\s+rate\b",
        r"\berror\s+rate\b",
        r"\bdrop\s+off\b",
        r"\bdropoff\b",
        r"\bcost\b",
        r"\bcac\b",
        r"\brisk\b",
        r"\bdefect\b",
        r"\bfriction\b",
        r"\bbilling\s+crm\s+delta\b",
        r"\bcrm\s+billing\s+delta\b",
        r"\bbilling\s+delta\b",
        r"\bcrm\s+delta\b",
        r"\bdiscrepanc(?:y|ies)\b",
    )
    positive_patterns = (
        r"\bactivation\b",
        r"\bconversion\b",
        r"\bqualified\b",
        r"\bretention\b",
        r"\brevenue\b",
        r"\barr\b",
        r"\bpipeline\b",
        r"\bwin\s+rate\b",
        r"\badoption\b",
        r"\bcompletion(?:\s+rate)?\b",
        r"\bresponse\s+rate\b",
        r"\bsatisfaction\b",
        r"\bquality\b",
        r"\bdata\s+quality\b",
        r"\bdq\s+score\b",
        r"\bcoverage\b",
    )
    if any(re.search(pattern, text, re.I) for pattern in inverse_patterns):
        return "down"
    if any(re.search(pattern, text, re.I) for pattern in positive_patterns):
        return "up"
    if re.search(r"\b(?:down|decrease|reduce|lower|below|under|fewer|less)\b", text, re.I):
        return "down"
    if re.search(r"\b(?:up|increase|improve|raise|above|over|more|higher)\b", text, re.I):
        return "up"
    return ""


def _expected_trend_text(direction: str) -> str:
    if direction in {"up", "down"}:
        return f"Expected trend under remediation: {direction}"
    return "Expected trend should be reviewed against the Decision Gates."


def _movement_against_direction_text(direction: str) -> str:
    if direction == "up":
        return "Flat or down against expected remediation trend"
    if direction == "down":
        return "Flat or up against expected remediation trend"
    return "Movement against the expected Decision Gate trend"


def _extract_concrete_monitoring_values(value: str) -> tuple[str, ...]:
    text = _text(value)
    number = r"\d+(?:\.\d+)?"
    value_boundary = r"(?=$|[\s,.;:)>\]|])"
    unit = r"(?:%|pp|h|hrs?|hours?|business\s+days?|days?|weeks?|months?)"
    patterns = [
        rf"(?:[<>]=?|≥|≤)\s*{number}\s*{unit}{value_boundary}",
        rf"(?:[<>]=?|≥|≤)\s*{number}{value_boundary}",
        rf"\bwithin\s+{number}\s*(?:hours?|hrs?|business\s+days?|days?|weeks?|months?)\b",
        rf"\bDay\s+{number}\b",
        rf"\b{number}[- ]day\s+rolling(?:\s+baseline)?\b",
        rf"\b{number}[- ]week\s+rolling(?:\s+baseline)?\b",
        rf"\b{number}\s*{unit}{value_boundary}",
    ]
    values: list[str] = []
    spans: list[tuple[int, int]] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.I):
            start, end = match.span()
            if any(start < existing_end and end > existing_start for existing_start, existing_end in spans):
                continue
            spans.append((start, end))
            values.append(match.group(0))
    return tuple(_unique(values))


def _concrete_threshold_or_placeholder(values: tuple[str, ...], fallback: str) -> str:
    if values:
        return "; ".join(values[:3])
    return fallback


def _first_cadence_value(values: tuple[str, ...]) -> str:
    for value in values:
        if re.search(r"\b(?:rolling|Day\s+\d+|within\s+\d)", value, re.I):
            return value
    return ""


def _canary_target_text(*, signal: str, direction: str, meaning: str) -> str:
    if meaning and direction:
        return f"{meaning}; {_expected_trend_text(direction)}"
    if meaning:
        return meaning
    if direction:
        return _expected_trend_text(direction)
    if signal:
        return "Expected trend should be reviewed against the Decision Gates."
    return ""


def _canary_threshold_text(window: str, direction: str, concrete_values: tuple[str, ...]) -> str:
    if window:
        trend = _movement_against_direction_text(direction) if direction else "Movement against expected trend"
        return f"{trend} over {window}"
    return _concrete_threshold_or_placeholder(concrete_values, THRESHOLD_NOT_CONFIRMED)


def _spreadsheet_safe_cell(value: Any, *, audience: str, header: str) -> str:
    text = _redact_basic(_text(value))
    if audience == "client":
        text = _client_safe_cell(text, header=header)
    text = re.sub(r"\s+", " ", text).strip()
    if text and text[0] in ("=", "+", "-", "@"):
        text = "'" + text
    return text


def _client_safe_cell(value: str, *, header: str) -> str:
    text = value
    if header == "Source / evidence source" and _has_raw_source_ref(text):
        if re.search(r"\bupload:", text, re.I):
            return "Uploaded project document"
        if re.search(r"\b(?:storage_ref|storage:|source_ref)\b", text, re.I):
            return EVIDENCE_SOURCE_UNAVAILABLE
    text = re.sub(r"\bupload:[^\s|,)>\]]+", "Uploaded project document", text, flags=re.I)
    text = re.sub(r"\bstorage_ref\s*[:=]\s*[^\s|,)>\]]+", EVIDENCE_SOURCE_UNAVAILABLE, text, flags=re.I)
    text = re.sub(r"\bstorage:[^\s|,)>\]]+", EVIDENCE_SOURCE_UNAVAILABLE, text, flags=re.I)
    text = re.sub(r"\bsource_ref\s*[:=]\s*[^\s|,)>\]]+", EVIDENCE_SOURCE_UNAVAILABLE, text, flags=re.I)
    text = re.sub(r"\bknowledge[_-][A-Za-z0-9_.:-]+\b", "project evidence", text, flags=re.I)
    text = re.sub(r"\b(?:ev|evidence)[-_][A-Za-z0-9_.:-]+\b", "project evidence", text, flags=re.I)
    text = re.sub(r"\bfile[-_][A-Za-z0-9_.:-]+\b", "project file", text, flags=re.I)
    text = re.sub(r"\b(?:BF|RPN|H_norm|rho|scenario_probability|diagnostic score|operator trace)\b[^;,.|]*", "internal diagnostic redacted", text, flags=re.I)
    text = re.sub(r"\b(?:row source|diagnostic notes?|internal source refs?|operator-only)\b", "internal detail", text, flags=re.I)
    text = _client_monitoring_language_polish(text)
    text = _redact_client_internal_metadata(text)
    return text


def _client_monitoring_language_polish(value: str) -> str:
    text = str(value or "")
    text = re.sub(r"\barchitecture hypothesis\b", "automation architecture assumption", text, flags=re.I)
    text = re.sub(r"\bH5\b|\bhypothesis\s+5\b", "technical feasibility check", text, flags=re.I)
    text = re.sub(r"\bH9\b|\bhypothesis\s+9\b", "momentum assumption", text, flags=re.I)
    return text


def _redact_client_internal_metadata(value: str) -> str:
    text = str(value or "")
    text = re.sub(r"\bpolicy_audit_log\b", "operator audit log redacted", text, flags=re.I)
    text = re.sub(r"\braw_provider_payload\b", "provider payload redacted", text, flags=re.I)
    text = re.sub(r"\braw[\s_-]+prompt\b", "prompt redacted", text, flags=re.I)
    text = re.sub(r"\bproject_state\s*\.\s*json\b", "internal project state redacted", text, flags=re.I)
    text = re.sub(r"\bmachine_archive\b", "internal machine archive redacted", text, flags=re.I)
    text = re.sub(
        r"\bruntime[\s_/-]*preflight(?:[\s_-]+metadata)?\b",
        "runtime metadata redacted",
        text,
        flags=re.I,
    )
    text = re.sub(r"\bupload_store\b", "internal upload storage", text, flags=re.I)
    return text


def _redact_basic(value: str) -> str:
    text = _text(value)
    text = re.sub(r"(?i)(api[_-]?key|token|password|secret|credential)\s*[:=]\s*[^,\s|;]+", r"\1=[REDACTED]", text)
    text = re.sub(r"(?i)\bBearer\s+[A-Za-z0-9._~+/=-]+", "Bearer [REDACTED]", text)
    text = re.sub(r"\bsk-[A-Za-z0-9_\-]{8,}\b", "[REDACTED]", text)
    text = re.sub(r"(?i)\b[A-Z]:[\\/][^\s|]+", "[REDACTED]", text)
    text = re.sub(r"\\\\[A-Za-z0-9_.-]+\\[^\s|]+", "[REDACTED]", text)
    text = re.sub(r"file://[^\s|]+", "[REDACTED]", text)
    text = re.sub(r"(?<!https:)(?<!http:)/(Users|home|mnt|var|tmp|root|workspace|Volumes|opt)/[^\s|]+", "[REDACTED]", text)
    return text


def _has_raw_source_ref(value: str) -> bool:
    return bool(re.search(r"\b(?:upload:|storage_ref|storage:|source_ref|knowledge[_-])", value or "", re.I))


def _extract_heading_section(markdown: str, heading: str) -> str:
    normalized_target = _normalize_label(heading)
    lines = str(markdown or "").splitlines()
    capture = False
    level = 0
    captured: list[str] = []
    for line in lines:
        match = re.match(r"^(#{1,6})\s+(.+?)\s*$", line.strip())
        if match:
            current_level = len(match.group(1))
            normalized = _normalize_label(match.group(2))
            if capture and current_level <= level:
                break
            if normalized == normalized_target:
                capture = True
                level = current_level
                continue
        if capture:
            captured.append(line)
    return "\n".join(captured).strip()


def _markdown_table_blocks(section: str) -> list[list[str]]:
    blocks: list[list[str]] = []
    current: list[str] = []
    for line in str(section or "").splitlines():
        if line.strip().startswith("|") and "|" in line.strip()[1:]:
            current.append(line.strip())
            continue
        if current:
            blocks.append(current)
            current = []
    if current:
        blocks.append(current)
    return blocks


def _split_table_row(line: str) -> list[str]:
    stripped = line.strip().strip("|")
    return [cell.strip() for cell in stripped.split("|")]


def _looks_like_separator(line: str) -> bool:
    return bool(re.match(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)+\|?\s*$", line or ""))


def _is_clear_decision_gate_header(headers: list[str]) -> bool:
    text = " ".join(headers)
    has_gate_context = any(term in text for term in ("gate", "decision", "signal", "metric", "criterion", "condition"))
    has_threshold_context = any(term in text for term in ("threshold", "stop", "proceed", "extend", "warning", "good", "target", "action"))
    return has_gate_context and has_threshold_context


def _first_value(cells: dict[str, str], keywords: tuple[str, ...]) -> str:
    for keyword in keywords:
        for header, value in cells.items():
            if keyword in header and _text(value):
                return value
    return ""


def _nth(values: list[str], index: int) -> str:
    return values[index] if index < len(values) else ""


def _clean_placeholder(value: Any, fallback: str) -> str:
    text = _text(value)
    if not text or text.strip("-") == "":
        return fallback
    if text.lower() in {"tbd", "n/a", "none", "unknown"}:
        return fallback
    return text


def _extract_ids(value: str, *, prefix: str) -> list[str]:
    pattern = rf"\b{re.escape(prefix)}\d+\b"
    return _unique(re.findall(pattern, value or "", re.I))


def _extract_evidence_ids(value: str) -> list[str]:
    return _unique(re.findall(r"\b(?:ev|evidence|knowledge)[-_][A-Za-z0-9_.:-]+\b", value or "", re.I))


def _extract_source_refs(value: str) -> list[str]:
    refs = re.findall(r"\b(?:upload|storage):[^\s|,)>\]]+", value or "", re.I)
    refs.extend(re.findall(r"\bsource_ref\s*[:=]\s*[^\s|,)>\]]+", value or "", re.I))
    return _unique(refs)


def _unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        text = _text(value)
        key = text.lower()
        if text and key not in seen:
            seen.add(key)
            output.append(text)
    return output


def _normalize_label(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def _enum_value(value: Any) -> str:
    return _text(getattr(value, "value", value))


def _text(value: Any) -> str:
    return str(value if value is not None else "").replace("\n", " ").replace("\r", " ").strip()
