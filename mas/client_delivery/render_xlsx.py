"""XLSX renderer for the Client Delivery Generator v0.5."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import DeliveryPackage
from .utils import safe_join, safe_text


SHEET_NAMES = [
    "Decision Summary",
    "30-60-90 Actions",
    "KPI Tracker",
    "Assumptions",
    "Review Triggers",
]


def render_execution_tracker_xlsx(package: DeliveryPackage, path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.active.title = SHEET_NAMES[0]
    for name in SHEET_NAMES[1:]:
        workbook.create_sheet(name)

    _render_decision_summary(workbook["Decision Summary"], package)
    _render_actions(workbook["30-60-90 Actions"], package)
    _render_kpis(workbook["KPI Tracker"], package)
    _render_assumptions(workbook["Assumptions"], package)
    _render_review_triggers(workbook["Review Triggers"], package)

    for worksheet in workbook.worksheets:
        _style_sheet(worksheet)

    workbook.save(output_path)
    return output_path


def _render_decision_summary(ws, package: DeliveryPackage) -> None:
    rows = [
        ["Field", "Value"],
        ["project_id", package.project_id],
        ["decision_statement", package.decision_statement],
        ["selected_option", package.recommendation.selected_option],
        ["recommendation_confidence", package.recommendation.confidence],
        ["evidence_strength", package.recommendation.evidence_strength],
        ["source_report_excerpt", package.source_report_excerpt],
    ]
    _append_rows(ws, rows)


def _render_actions(ws, package: DeliveryPackage) -> None:
    rows = [["phase", "action", "owner", "dependencies", "evidence", "notes", "status", "success_criteria"]]
    for action in package.execution_plan:
        rows.append(
            [
                action.phase,
                action.action,
                action.owner,
                safe_join(action.dependencies),
                safe_join(action.evidence),
                action.notes,
                action.status,
                action.success_criteria,
            ]
        )
    _append_rows(ws, rows)


def _render_kpis(ws, package: DeliveryPackage) -> None:
    rows = [["name", "indicator_type", "threshold_red", "threshold_amber", "actual_value", "status", "owner", "cadence", "notes"]]
    numeric_threshold_rows: list[tuple[int, float, float]] = []
    for row_idx, kpi in enumerate(package.kpis, start=2):
        red = _parse_number(kpi.threshold_red)
        amber = _parse_number(kpi.threshold_amber)
        if red is not None and amber is not None:
            numeric_threshold_rows.append((row_idx, red, amber))
        rows.append(
            [
                kpi.name,
                kpi.indicator_type,
                _numeric_or_text(kpi.threshold_red),
                _numeric_or_text(kpi.threshold_amber),
                _numeric_or_text(kpi.actual_value),
                kpi.status,
                kpi.owner,
                kpi.cadence,
                kpi.notes,
            ]
        )
    _append_rows(ws, rows)
    _apply_kpi_conditional_formatting(ws, numeric_threshold_rows)


def _render_assumptions(ws, package: DeliveryPackage) -> None:
    rows = [["assumption", "falsification_trigger", "owner", "confidence", "evidence", "notes"]]
    for assumption in package.critical_assumptions:
        rows.append(
            [
                assumption.assumption,
                assumption.falsification_trigger,
                assumption.owner,
                assumption.confidence,
                safe_join(assumption.evidence),
                assumption.notes,
            ]
        )
    _append_rows(ws, rows)


def _render_review_triggers(ws, package: DeliveryPackage) -> None:
    rows = [["cadence", "owner", "trigger", "notes"]]
    for trigger in package.review.reentry_triggers:
        rows.append([package.review.cadence, package.review.owner, safe_text(trigger), package.review.notes])
    _append_rows(ws, rows)


def _append_rows(ws, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append([value if isinstance(value, (int, float)) and not isinstance(value, bool) else safe_text(value) for value in row])


def _style_sheet(ws) -> None:
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = header_font
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(safe_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 14), 42)
    ws.freeze_panes = "A2"


def _apply_kpi_conditional_formatting(ws, numeric_threshold_rows: list[tuple[int, float, float]]) -> None:
    red_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
    amber_fill = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    for row_idx, red, amber in numeric_threshold_rows:
        actual_cell = f"E{row_idx}"
        ws.conditional_formatting.add(
            actual_cell,
            CellIsRule(operator="lessThanOrEqual", formula=[str(red)], fill=red_fill, stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            actual_cell,
            CellIsRule(operator="lessThanOrEqual", formula=[str(amber)], fill=amber_fill),
        )


def _numeric_or_text(value: Any) -> float | str:
    parsed = _parse_number(value)
    if parsed is not None:
        return parsed
    return safe_text(value)


def _parse_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = safe_text(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1].strip()
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return None
    try:
        return float(text)
    except ValueError:
        return None
